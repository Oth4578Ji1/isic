import json
import traceback
from django.contrib import messages
from venv import logger
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect, HttpResponseServerError, JsonResponse, HttpResponse
from .forms import ProfileUpdateForm, SignUpForm, SignInForm
from .models import Application, Competition, CustomUser
from django.db.models import Count
from openpyxl import Workbook
from django.views.decorators.csrf import csrf_exempt
import json
from django.views.decorators.http import require_http_methods
from isicApp import models
from django.utils.timesince import timesince
from django.utils import timezone
from django.shortcuts import render
from django.http import JsonResponse
from .models import Competition
from datetime import datetime, time
import logging
from django.db import connection
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
import logging


    

def home(request):
    return render(request, 'base.html')


def is_etudiant(user):
    return user.user_type == 'etudiant'

# is responsable
def is_responsable(user):
    return user.user_type == 'responsable'



@login_required(login_url='sign-in')
def user_profile(request):
    if request.user.user_type == 'etudiant':
        # Fetch the student's latest application
        latest_application = Application.objects.filter(student=request.user).order_by('-applied_at').first()
        
        # Prepare the context with application status
        context = {
            'latest_application': latest_application,
            'status_colors': {
                'Arrivée': '#0044CD',
                'En cours': '#ffc107',
                'Accepté': '#4caf50',
                'Rejetée': '#f44336'
            }
        }
        return render(request, 'edu-dash.html', context)
    elif request.user.user_type == 'responsable':
        return render(request, 'res-dash.html')
    else:
        return redirect('sign-in')



def sign_up(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.user_type = 'etudiant'
            user.save()
            login(request, user)
            return JsonResponse({'success': True, 'redirect_url': '/edu-dash/'})
        else:
            errors = []
            for field in form:
                for error in field.errors:
                    errors.append(f"{field.label}: {error}")
            return JsonResponse({'success': False, 'errors': errors})
    else:
        form = SignUpForm()
    return render(request, 'sign-up.html', {'form': form})


def sign_in(request):
    if request.method == 'POST':
        form = SignInForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                if user.is_staff:
                    return JsonResponse({'success': True, 'redirect_url': '/res-dash/'})
                else:
                    return JsonResponse({'success': True, 'redirect_url': '/edu-dash/'})
            else:
                return JsonResponse({'success': False, 'errors': ['Invalid username or password']})
        else:
            errors = [f"{field}: {error[0]}" for field, error in form.errors.items()]
            return JsonResponse({'success': False, 'errors': errors})
    else:
        form = SignInForm()
    return render(request, 'sign-in.html', {'form': form})

@login_required(login_url='sign-in')
def sign_out(request):
    logout(request)
    return redirect('sign-in')



@login_required(login_url='sign-in')
@user_passes_test(is_etudiant, login_url='sign-in')
def profile_view(request):
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'Invalid form data'})
    else:
        form = ProfileUpdateForm(instance=request.user)
    return render(request, 'profile.html', {'form': form})



@login_required(login_url='sign-in')
@user_passes_test(is_responsable, login_url='sign-in')
def view_application(request):
    item=Application.objects.all()
    context={
        'item':item
    }
    return render(request, 'applications.html', context)




@login_required(login_url='sign-in')
@user_passes_test(is_responsable, login_url='sign-in')
def res_dash(request):
    # Count pending and processed applications
    pending_applications = Application.objects.filter(status='Arrivée').count()
    processed_applications = Application.objects.exclude(status='Arrivée').count()

    # Get application types for pie chart
    application_types = Application.objects.values('competition__type').annotate(count=Count('id'))
    
    # Get all applications for the table
    all_applications = Application.objects.select_related('student', 'competition').all()

    application_types = list(Application.objects.values('competition__type').annotate(count=Count('id')))

    context = {
        'pending_applications': pending_applications,
        'processed_applications': processed_applications,
        'application_types': json.dumps(application_types),
        'all_applications': all_applications,
    }
    return render(request, 'res-dash.html', context)



def export_excel(request):
    logger = logging.getLogger(__name__)
    
    type_filter = request.GET.get('type')
    field_filter = request.GET.get('field')
    status_filter = request.GET.get('status')
    language_filter = request.GET.get('language')
    
    logger.info(f"Received filters: type={type_filter}, field={field_filter}, status={status_filter}, language={language_filter}")
    
    applications = Application.objects.all()
    logger.info(f"Initial query count: {applications.count()}")
    
    logger.info(f"Final query count: {applications.count()}")
    
    # Log the SQL query
    logger.info(f"SQL Query: {applications.query}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    
    # Define header styles
    header_font = Font(bold=False, color="FFFFFF")
    header_fill = PatternFill(start_color="0044CD", end_color="0044CD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    headers = [
        'Nom', 'Prenom', 'Email', 'CIN', 'Adress', 'Date Naissance', 
        'Lieu de naissance', 'N° Telephone', 'CNE', 'Type de BAC', 
        'Note de Bac', 'Type diplome', 'Anne de diplome', 'Mention diplome', 
        'Filiere License', 'Universite', '3eme Langue', 'Note S1', 
        'Note S2', 'Note S3', 'Note S4', 'Note S5', 'Note S6', 
        'Type de Concours', 'Filiere', 'Langue de concours', 'Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[cell.column_letter].width = len(header) + 30  # Create space between headers
    
    # Add data to the sheet
    for app in applications:
        ws.append([
            app.student.first_name,
            app.student.last_name,
            app.student.email,
            app.student.CIN,
            app.student.address,
            app.student.date_of_birth,
            app.student.place_of_birth,
            app.student.phone,
            app.CNE,
            app.baccalaureate_category,
            app.baccalaureate_grade,
            app.diploma_category,
            app.diploma_year,
            app.diploma_grade,
            app.license_field_name,
            app.université,
            app.troisieme_lng,
            app.semester_1_grade,
            app.semester_2_grade,
            app.semester_3_grade,
            app.semester_4_grade,
            app.semester_5_grade,
            app.semester_6_grade,
            app.competition.get_type_display(),
            app.competition.field,
            app.language,
            app.status,
        ])
    
    # Create HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=applications.xlsx'
    
    # Save the workbook to the response
    wb.save(response)
    
    # Log all SQL queries executed
    for query in connection.queries:
        logger.info(f"Executed SQL: {query['sql']}")
    
    return response


def export_pdf(request):
    logger = logging.getLogger(__name__)

    type_filter = request.GET.get('type')
    field_filter = request.GET.get('field')
    status_filter = request.GET.get('status')
    language_filter = request.GET.get('language')
    
    logger.info(f"Received filters: type={type_filter}, field={field_filter}, status={status_filter}, language={language_filter}")
    
    # Utilisation de filter() au lieu de all() pour appliquer les filtres
    applications = Application.objects.filter(status='Accepté')
    logger.info(f"Initial query count: {applications.count()}")
    
    # Appliquer des filtres supplémentaires si nécessaire
    if type_filter:
        applications = applications.filter(competition__type=type_filter)
    if field_filter:
        applications = applications.filter(competition__field=field_filter)
    if language_filter:
        applications = applications.filter(language=language_filter)
    
    logger.info(f"Final query count: {applications.count()}")
    
    # Log the SQL query
    logger.info(f"SQL Query: {applications.query}")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=applications.pdf'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Define the style for the table
    styles = getSampleStyleSheet()
    header_style = styles['Heading4']
    body_style = styles['BodyText']
    
    # Define headers
    headers = [
        'Nom', 'Prenom', 'CNE', 'Type de Concours', 'Filiere', 'Langue de concours'
    ]
    
    # Prepare data for the table
    data = [headers]
    
    for app in applications:
        data.append([
            app.student.last_name,
            app.student.first_name,
            app.CNE,
            app.competition.get_type_display(),
            app.competition.field,
            app.language,
        ])
    
    # Create the table
    table = Table(data)
    
    # Define table style
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), '#0044CD'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Log all SQL queries executed
    for query in connection.queries:
        logger.info(f"Executed SQL: {query['sql']}")
    
    return response




@login_required(login_url='sign-in')
@user_passes_test(is_responsable, login_url='sign-in')
def edit_application(request, application_id):
    application = get_object_or_404(Application, id=application_id)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in ['Arrivée', 'En cours', 'Accepté', 'Rejetée']:
            application.status = new_status
            application.save()
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'Invalid status selected.'})
    
    context = {
        'application': application,
        'status_choices': ['Arrivée', 'En cours', 'Accepté', 'Rejetée']
    }
    return render(request, 'edit.html', context)

@login_required(login_url='sign-in')
@user_passes_test(is_responsable, login_url='sign-in')
def view_app(request, application_id):
    application = get_object_or_404(Application, id=application_id)
    context = {
        'application': application,
    }
    return render(request, 'view.html', context)





@login_required(login_url='sign-in')
@user_passes_test(is_responsable, login_url='sign-in')
def add_competition(request):
    return render(request, 'add-competition.html')



logger = logging.getLogger(__name__)

@csrf_exempt
@require_http_methods(["POST"])
def api_add_competition(request):
    try:
        type = request.POST.get('type')
        field = request.POST.get('field')
        deadline = request.POST.get('deadline')

        if not type or not deadline:
            return JsonResponse({'success': False, 'error': 'Type and deadline are required'}, status=400)

        # Map the full words to single-letter codes
        type_map = {'License': 'L', 'Master': 'M'}
        type_code = type_map.get(type)
        if not type_code:
            return JsonResponse({'success': False, 'error': 'Invalid competition type'}, status=400)

        competition = Competition.objects.create(
            type=type_code,
            field=field,
            deadline=deadline
        )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    


def get_competitions(request):
    competitions = Competition.objects.all().values('type', 'field', 'deadline', 'created_at')
    return JsonResponse(list(competitions), safe=False)

def get_application_stats(request):
    stats = Application.objects.values('competition__field').annotate(count=Count('id')).order_by('competition__field')
    return JsonResponse(list(stats), safe=False)

@login_required(login_url='sign-in')
@user_passes_test(is_responsable, login_url='sign-in')
def overview(request):
    return render(request, 'overview.html')


@login_required(login_url='sign-in')
def latest_competitions(request):
    competitions = Competition.objects.order_by('-created_at')[:5]
    data = [
        {
            'id': comp.id,
            'type': comp.get_type_display(),
            'field': comp.field,
            'deadline': comp.deadline.strftime('%d-%m-%Y'),
            'created_at': comp.created_at,
        }
        for comp in competitions
    ]
    return JsonResponse(data, safe=False)


@login_required(login_url='sign-in')
@user_passes_test(is_etudiant, login_url='sign-in')
@require_http_methods(["GET", "POST"])
def add_application(request, competition_id):
    competition = get_object_or_404(Competition, id=competition_id)
    
    if request.method == 'POST':
        try:
            # Create a new application
            application = Application(
                student=request.user,
                competition=competition,
                CNE=request.POST.get('CNE'),
                baccalaureate_category=request.POST.get('baccalaureate_category'),
                baccalaureate_grade=float(request.POST.get('baccalaureate_grade')),
                language=request.POST.get('language'),
                # Add other fields as necessary
            )
            
            # If it's a Master's application, add the additional fields
            if competition.type == 'M':
                application.diploma_category = request.POST.get('diploma_category')
                application.diploma_year = int(request.POST.get('diploma_year'))
                application.diploma_grade = request.POST.get('diploma_grade')
                application.semester_1_grade = float(request.POST.get('semester_1_grade'))
                application.semester_2_grade = float(request.POST.get('semester_2_grade'))
                application.semester_3_grade = float(request.POST.get('semester_3_grade'))
                application.semester_4_grade = float(request.POST.get('semester_4_grade'))
                application.semester_5_grade = float(request.POST.get('semester_5_grade'))
                application.semester_6_grade = float(request.POST.get('semester_6_grade'))
                application.license_field_name = request.POST.get('license_field_name')
                application.université = request.POST.get('université')
            
            application.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    context = {
        'competition': competition
    }
    return render(request, 'add-application.html', context)

@login_required
def get_user_info(request):
    user = request.user
    return JsonResponse({
        'first_name': user.first_name,
        'last_name': user.last_name,
        'address': user.address,
        'date_of_birth': user.date_of_birth.strftime('%Y-%m-%d') if user.date_of_birth else '',
        'email': user.email,
        'phone': user.phone,
        'CIN': user.CIN,
        'place_of_birth': user.place_of_birth,
    })

def submit_general_application(request):
    return render(request, 'add-app-normal.html')


def my_apps_view(request):
    return render(request, 'my-apps.html')


@login_required
@require_http_methods(["GET"])
def user_applications(request):
    applications = Application.objects.filter(student=request.user).order_by('-applied_at')
    data = []
    for app in applications:
        data.append({
            'id': app.id,
            'competition_type': app.competition.get_type_display(),
            'competition_field': app.competition.field,
            'status': app.status,
            'submitted_date': app.applied_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    return JsonResponse(data, safe=False)

@login_required
@require_http_methods(["GET"])
def application_details(request, application_id):
    application = get_object_or_404(Application, id=application_id, student=request.user)
    data = {
        'id': application.id,
        'competition_type': application.competition.get_type_display(),
        'competition_field': application.competition.field,
        'status': application.status,
        'CNE': application.CNE,
        'baccalaureate_category': application.baccalaureate_category,
        'baccalaureate_grade': application.baccalaureate_grade,
        'language': application.language,
        'competition_deadline': application.competition.deadline.strftime('%Y-%m-%d'),
    }
    
    if application.competition.type == 'M':
        data.update({
            'diploma_category': application.diploma_category,
            'diploma_year': application.diploma_year,
            'diploma_grade': application.diploma_grade,
            'semester_1_grade': application.semester_1_grade,
            'semester_2_grade': application.semester_2_grade,
            'semester_3_grade': application.semester_3_grade,
            'semester_4_grade': application.semester_4_grade,
            'semester_5_grade': application.semester_5_grade,
            'semester_6_grade': application.semester_6_grade,
            'license_field_name': application.license_field_name,
            'université': application.université,
        })
    
    return JsonResponse(data)

logger = logging.getLogger(__name__)

@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_application(request, application_id):
    try:
        application = get_object_or_404(Application, id=application_id, student=request.user)
        
        # Convert the deadline date to a datetime at the end of the day
        deadline_datetime = timezone.make_aware(
            datetime.combine(application.competition.deadline, time(23, 59, 59))
        )
        
        if timezone.now() > deadline_datetime:
            return JsonResponse({'success': False, 'error': 'La date limite de modification est dépassée.'})
        
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError as e:
            logger.error(f"JSON Decode Error: {str(e)}")
            return JsonResponse({'success': False, 'error': f'Invalid JSON data: {str(e)}'}, status=400)
        
        # Log received data for debugging
        logger.debug(f"Received data: {data}")
        
        # Convert the array of objects to a dictionary if necessary
        if isinstance(data, list):
            data_dict = {item['name']: item['value'] for item in data}
        else:
            data_dict = data
        
        # Update the fields
        for field, value in data_dict.items():
            if hasattr(application, field):
                setattr(application, field, value)
            else:
                logger.warning(f"Attempted to set non-existent field: {field}")
        
        application.save()
        return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"Error in update_application: {str(e)}\n{traceback.format_exc()}")
        return HttpResponseServerError(json.dumps({'success': False, 'error': str(e)}), content_type="application/json")
    

def application_details_view(request, application_id):
    application = get_object_or_404(Application, id=application_id, student=request.user)
    context = {
        'application': application,
    }
    return render(request, 'application_details.html', context)


@login_required(login_url='sign-in')
def notifications_view(request):
    return render(request, 'notifications-edu.html')


@login_required(login_url='sign-in')
def latest_applications(request):
    applications = Application.objects.order_by('-applied_at')[:5]
    data = [
        {
            'id': apps.id,
            'nom': apps.student.last_name,
            'prenom': apps.student.first_name,
            'competition_type': apps.competition.get_type_display(),
            'competition_field': apps.competition.field,
            'competition_deadline': apps.competition.deadline,
            'applied_at': apps.applied_at,
            'status': apps.status,
        }
        for apps in applications
    ]
    return JsonResponse(data, safe=False)

@login_required(login_url='sign-in')
def notifications_res_view(request):
    return render(request, 'notifications-res.html')